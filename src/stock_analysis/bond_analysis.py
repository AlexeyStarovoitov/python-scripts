import pandas as pd
import numpy as np
import argparse
import re
from enum import Enum
from openpyxl.reader import excel
from openpyxl.worksheet import dimensions
# Data imported from snowball income

class ColumnName(Enum):
    ISIN=1
    PURCHASE_PRICE=2
    NOMINAL_PRICE=3
    REMAIN_NOMINAL_PRICE=11
    COUPON=4
    COUPON_FREQUENCY=5
    PURCHASE_DATE=6
    EXPIRATION_DATE=7
    COUPON_YIELD=8
    YIELD=9
    RATING=10

class ColumnNameMapperColName(Enum):
    COLUMN_ENUM=1
    COLUMN_TEMPLATE=2

class ColumnNameMapper:
    def __init__(self):
        col_data = [ [ColumnName.ISIN, ['^актив']],
                     [ColumnName.PURCHASE_PRICE, ['^Средняя цена']],
                     [ColumnName.NOMINAL_PRICE, ['^Номинал']],
                     [ColumnName.REMAIN_NOMINAL_PRICE, ['^Остаточный номинал']],
                     [ColumnName.COUPON, ['^Купонная доходность']],
                     [ColumnName.COUPON_FREQUENCY, ['^Частота купона']],
                     [ColumnName.PURCHASE_DATE, ['^Дата покупки']],
                     [ColumnName.EXPIRATION_DATE, ['^Дата погашения']],
                     [ColumnName.COUPON_YIELD, ['^Купонная доходность']],
                     [ColumnName.YIELD, ['^Доходность']],
                     [ColumnName.RATING, ['^Надежность']]
        ]
        
        self._map_data = pd.DataFrame(data = col_data, columns = [ColumnNameMapperColName.COLUMN_ENUM, 
                                                                            ColumnNameMapperColName.COLUMN_TEMPLATE])
    @staticmethod
    def _check_template_match(col_name, tem_list):
        for temp in tem_list:
            if re.search(temp, col_name, re.IGNORECASE):
                return True
        else:
            return False
    def map_column(self, col_name):
        for i in self._map_data.index:
            cur_name = self._map_data.loc[i, :]
            if ColumnNameMapper._check_template_match(col_name, cur_name[ColumnNameMapperColName.COLUMN_TEMPLATE]):
                return cur_name[ColumnNameMapperColName.COLUMN_ENUM]
        else:
            return None
    def map_columns(self, columns):
        map_columns = [self.map_column(column) for column in columns]
        return map_columns

class BondConverter:
    _column_width = 15
    def __init__(self, csv_path):
        self._bond_db = BondConverter._bond_db_init(csv_path)
    @staticmethod
    def _bond_db_init(csv_path):
        _bond_db = pd.read_csv(csv_path)
        _bond_db.dropna(inplace=True, axis=1)
        return _bond_db
    def bond_convert_to_excel(self, excel_path):
        writer = pd.ExcelWriter(path=excel_path, engine='openpyxl')
        self._bond_db.to_excel(excel_writer=writer, float_format="%.2f")
        writer.save()

        wb = excel.load_workbook(excel_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]    
            for row in ws.iter_rows(min_col=1, max_col=ws.max_column, max_row=1):
                for col in row:
                    ws.column_dimensions[col.column_letter] = dimensions.ColumnDimension(ws, index=col.column_letter, width=self._column_width)

        wb.save(excel_path)

if __name__=='__main__':
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("-csv_db", dest="csv_db", type=str)
    arg_parser.add_argument("-out_file", dest="out_file", type=str)
    args = arg_parser.parse_args()

    bond_converter = BondConverter(args.csv_db)
    bond_converter.bond_convert_to_excel(args.out_file)
    