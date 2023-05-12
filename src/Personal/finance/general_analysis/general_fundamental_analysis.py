import pandas as pd
import numpy as np
import argparse
import re
from enum import Enum, Flag
from openpyxl import workbook, worksheet
from openpyxl.reader import excel
from openpyxl.worksheet import dimensions


class ColumnName(Enum):
    NAME=1
    CAPITALIZATION=2
    P_E=3
    P_BV=4
    P_S=14
    P_FCF=5
    EV_EBIT=6
    EV_S=7
    ROE=8
    TARGET_PE=9
    TARGET_P_FCF=10
    TARGET_EV_EBIT=11
    TARGET_P_S=12
    TARGET_NET_ASSET=13
    TARGET_DIV=17
    RES_TARGET=18
    D_E=15
    DIV_YIELD=16
    PRICE = 19
    
    
class ColumnNameMapperColName(Enum):
    COLUMN_ENUM=1
    COLUMN_TEMPLATE=2
    COLUMN_ADDUCED_NAME=3

class ANALYSIS_TYPE(Flag):
    NONE_ANALYSYS_TYPE=0x00
    PE_COMP_ANALYSYS_TYPE=0x01
    P_FCF_COMP_ANALYSIS_TYPE=0x02
    EV_EBIT_COM_ANALYSIS_TYPE=0x04
    P_S_COM_ANALYSIS_TYPE=0x08
    NET_ASSET_ANALYSIS_TYPE=0x10
    DIV_ANALYSIS_TYPE=0x20
    ALL_ANALYSIS_TYPE=0x3f

class ColumnNameOperation(Enum):
    COLUMN_MAP=1
    COLUMN_CONVERT=2

class ColumnNameMapper:
   
    def __init__(self):
        col_data = [ [ColumnName.NAME, ['^название[a-z]*[1-9]*', '^name[a-z]*[1-9]*'], 'name'],
                     [ColumnName.CAPITALIZATION, ['^капитализация[a-z]*[1-9]*'], 'cap'],
                     [ColumnName.P_E, ['^p/e[a-z]*[1-9]*'], 'P/E'],
                     [ColumnName.P_BV, ['^p/bv[a-z]*[1-9]*'], 'P/BV'],
                     [ColumnName.P_FCF, ['^p/fcf[a-z]*[1-9]*'], 'P/FCF'],
                     [ColumnName.P_S, ['^p/s[a-z]*[1-9]*'], 'P/S'],
                     [ColumnName.EV_EBIT, ['^ev/ebit[a-z]*[1-9]*'], 'EV/EBIT'],
                     [ColumnName.EV_S, ['^ev/s[a-z]*[1-9]*'], 'EV/S'],
                     [ColumnName.D_E, ['^debt/equity[a-z]*[1-9]*'], 'D/E'],
                     [ColumnName.ROE, ['^roe[a-z]*[1-9]*'], 'ROE'],
                     [ColumnName.DIV_YIELD, ['^тек[. ]*дох-ть[a-z]*[1-9]*'], 'Div. yield'],
                     [ColumnName.PRICE, ["^price[a-z]*[1-9]*"], "Price"],
                     [ColumnName.TARGET_PE, [], 'Target (P/E)'],
                     [ColumnName.TARGET_P_FCF, [], 'Target (P/FCF)'],
                     [ColumnName.TARGET_EV_EBIT, [], 'Target (EV/EBIT)'],
                     [ColumnName.TARGET_P_S, [], 'Target (P/S)'],
                     [ColumnName.TARGET_NET_ASSET, [], 'Target (P/BV)'],
                     [ColumnName.TARGET_DIV, [], 'Target (Div.)'],
                     [ColumnName.RES_TARGET, [], 'Res Target']
        ]
        
        self._map_data = pd.DataFrame(data = col_data, columns = [ColumnNameMapperColName.COLUMN_ENUM, 
                                                                            ColumnNameMapperColName.COLUMN_TEMPLATE, 
                                                                            ColumnNameMapperColName.COLUMN_ADDUCED_NAME])
    @staticmethod
    def _check_template_match(col_name, tem_list):
        for temp in tem_list:
            if re.search(temp, col_name, re.IGNORECASE):
                return True
        else:
            return False
    def map_column(self, col_name):
        for i in self._map_data.index:
            cur_name = self._map_data.loc[i, :]
            if ColumnNameMapper._check_template_match(col_name, cur_name[ColumnNameMapperColName.COLUMN_TEMPLATE]):
                return cur_name[ColumnNameMapperColName.COLUMN_ENUM]
        else:
            return None
    def map_columns(self, columns):
        map_columns = [self.map_column(column) for column in columns]
        return map_columns
    def enum2strconvert(self, column_enum):
        column_entry = self._map_data[self._map_data[ColumnNameMapperColName.COLUMN_ENUM]==column_enum].iloc[0,:]
        if not column_entry.empty:
            return column_entry[ColumnNameMapperColName.COLUMN_ADDUCED_NAME]
        return None
    def enum2strconvert_columns(self, column_enums):
        converted_columns = [self.enum2strconvert(column) for column in column_enums]
        return converted_columns
        

class GeneralFundamentalAnalysis:
    _p_bv = 1.5
    _column_width = 15
    _div_threashold = 10 # in percent
    def __init__(self, csv_db, analysis_type):
        self._db = pd.read_csv(csv_db)
        self._column_mapper = ColumnNameMapper()
        self._analysis_type = analysis_type
        self._preprocess()
    def _map_column_general_operation(self, op_enum):
        db = self._db
        columns = list(db.columns)
        if op_enum == ColumnNameOperation.COLUMN_MAP:
            map_columns = self._column_mapper.map_columns(columns)
        elif op_enum == ColumnNameOperation.COLUMN_CONVERT:
            map_columns = self._column_mapper.enum2strconvert_columns(columns)
        column_rename_dict = dict(zip(columns, map_columns))
        db.rename(columns=column_rename_dict, inplace=True)
        
    def _map_columns(self):
       self._map_column_general_operation(ColumnNameOperation.COLUMN_MAP)
    def _enum2str_convert_columns(self):
        self._map_column_general_operation(ColumnNameOperation.COLUMN_CONVERT)
    def _reorder_columns(self):
        db = self._db
        columns = list(db.columns)
        columns.remove(ColumnName.PRICE)
        columns.insert(len(columns), ColumnName.PRICE)
        db = db[columns]
        self._db = db
    def _add_analysis_columns(self):
        db = self._db
        _analysis_columns = []
        analysis_columns_mapper = {
            ANALYSIS_TYPE.PE_COMP_ANALYSYS_TYPE: ColumnName.TARGET_PE,
            ANALYSIS_TYPE.EV_EBIT_COM_ANALYSIS_TYPE: ColumnName.TARGET_EV_EBIT,
            ANALYSIS_TYPE.P_FCF_COMP_ANALYSIS_TYPE: ColumnName.TARGET_P_FCF,
            ANALYSIS_TYPE.P_S_COM_ANALYSIS_TYPE: ColumnName.TARGET_P_S,
            ANALYSIS_TYPE.NET_ASSET_ANALYSIS_TYPE: ColumnName.TARGET_NET_ASSET,
            ANALYSIS_TYPE.DIV_ANALYSIS_TYPE: ColumnName.TARGET_DIV
        }
        for (anal_mask, anal_col_name) in analysis_columns_mapper.items():
            if self._analysis_type & anal_mask:
                db[anal_col_name] = pd.Series([0]*len(db),index=db.index, dtype=np.float32)
                _analysis_columns.append(anal_col_name)
        
        db[ColumnName.RES_TARGET] = pd.Series([0]*len(db),index=db.index, dtype=np.float32)
        self._analysis_columns = _analysis_columns
    def _preprocess(self):
        self._map_columns()
        self._reorder_columns()
        self._add_analysis_columns()
        self._db.fillna(value = 0, inplace=True)
    @staticmethod
    def _map_proccessed_db(db, pr_db, proccess_column_name):
        for i, name in zip(pr_db[ColumnName.NAME].index, pr_db[ColumnName.NAME].to_numpy()):
            entry_index = db[ColumnName.NAME][db[ColumnName.NAME]==name].index[0]
            db.loc[entry_index, proccess_column_name] = pr_db.loc[i, proccess_column_name]
    
    def _calculate_via_general_analysis(self, parameter_column_name, result_column_name):
        db = self._db
        if parameter_column_name not in db.columns:
            return
        db_pr = db[db[parameter_column_name] > 0]
        if db_pr.empty:
            return
        means = db_pr.mean()
        #growth in percent
        #db_pr[result_column_name] = (means[parameter_column_name]-db_pr[parameter_column_name])/db_pr[parameter_column_name]*100 
        relative_target = (means[parameter_column_name]-db_pr[parameter_column_name])/db_pr[parameter_column_name]
        db_pr[result_column_name] = db_pr[ColumnName.PRICE]*(1 + relative_target)
        GeneralFundamentalAnalysis._map_proccessed_db(db, db_pr, result_column_name)
    def _calculate_via_comparative_pe_analysis(self):
        self._calculate_via_general_analysis(ColumnName.P_E, ColumnName.TARGET_PE)
        '''
        db_pe= db_pe[(db_pe[ColumnName.P_E] > 0) & 
                (db_pe[ColumnName.P_E] < means[ColumnName.P_E]) & 
                (db_pe[ColumnName.COM_ANALYSIS_EV_EBIT] > 0) &
                (db_pe[ColumnName.COM_ANALYSIS_EV_EBIT] < means[ColumnName.COM_ANALYSIS_EV_EBIT]) & 
                (db_pe[ColumnName.COM_ANALYSIS_EV_EBIT] < db_pe[ColumnName.P_E])] #& 
                #(db_pe[ColumnName.ROE] > means[ColumnName.ROE])]
        '''
    def _calculate_via_comparative_p_fcf_analysis(self):
        self._calculate_via_general_analysis(ColumnName.P_FCF, ColumnName.TARGET_P_FCF)
    def _calculate_via_comparative_ps_analysis(self):
        self._calculate_via_general_analysis(ColumnName.P_S, ColumnName.TARGET_P_S)
    def _calculate_via_comparative_ev_ebit(self):
        db = self._db
        for col_name in [ColumnName.EV_EBIT, ColumnName.EV_S, ColumnName.P_S, ColumnName.CAPITALIZATION]:
            if col_name not in db.columns:
                return
        db_pr = db[db[ColumnName.EV_EBIT] > 0]
        if db_pr.empty:
            return
        s = db_pr[ColumnName.CAPITALIZATION]/db_pr[ColumnName.P_S]
        ev = db[ColumnName.EV_S]*s
        ebit = ev/db[ColumnName.EV_EBIT]
        debt = ev - db_pr[ColumnName.CAPITALIZATION]
        means = db_pr.mean()
        fair_price =  means[ColumnName.EV_EBIT]*ebit-debt
        relative_target = (fair_price-db_pr[ColumnName.CAPITALIZATION])/fair_price
        db_pr[ColumnName.TARGET_EV_EBIT] = db_pr[ColumnName.PRICE]*(1 + relative_target)
        GeneralFundamentalAnalysis._map_proccessed_db(db, db_pr, ColumnName.TARGET_EV_EBIT)
    def _calculate_via_div_yeild(self):
        db = self._db
        #db_pr = db[db[ColumnName.DIV_YIELD] > self._div_threashold]
        if ColumnName.DIV_YIELD not in db.columns:
            return
        db_pr = db[db[ColumnName.DIV_YIELD] > 0]
        means = db_pr.mean()
        db_pr = db_pr[db_pr[ColumnName.DIV_YIELD] > means[ColumnName.DIV_YIELD]]
        db_pr[ColumnName.TARGET_DIV] = db_pr[ColumnName.PRICE]*(1 + db_pr[ColumnName.DIV_YIELD]/100)
        GeneralFundamentalAnalysis._map_proccessed_db(db, db_pr, ColumnName.TARGET_DIV)
    def _calcultate_result_price(self):
        db = self._db
        analysis_columns = self._analysis_columns
        for i in db.index:
            cur_entry = db.loc[i,:]
            target_prices = cur_entry[analysis_columns]
            target_prices = target_prices[target_prices > 0]
            db.loc[i,ColumnName.RES_TARGET] = target_prices.mean() if not target_prices.empty else 0
            #cur_entry.loc[ColumnName.RES_TARGET] = target_prices.mean()
            pass
        pass
    def _calculate_via_net_assets(self):
        db = self._db
        db_pr = db[db[ColumnName.P_BV] > 0]
        db_pr[ColumnName.NET_ASSET_ANALYSIS] = (self._p_bv-db_pr[ColumnName.P_BV])/self._p_bv*100
        GeneralFundamentalAnalysis._map_proccessed_db(db, db_pr, ColumnName.TARGET_NET_ASSET)
    def dump_to_excel(self, out_file_path):
        self._enum2str_convert_columns()
        writer = pd.ExcelWriter(path=out_file_path, engine='openpyxl')
        self._db.to_excel(excel_writer=writer, float_format="%.2f")
        writer.save()

        wb = excel.load_workbook(out_file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]    
            for row in ws.iter_rows(min_col=1, max_col=ws.max_column, max_row=1):
                for col in row:
                    ws.column_dimensions[col.column_letter] = dimensions.ColumnDimension(ws, index=col.column_letter, width=self._column_width)

        wb.save(out_file_path)
        
    def process(self):
        analysis_function_mapper = {
            ANALYSIS_TYPE.PE_COMP_ANALYSYS_TYPE: self._calculate_via_comparative_pe_analysis,
            ANALYSIS_TYPE.EV_EBIT_COM_ANALYSIS_TYPE: self._calculate_via_comparative_ev_ebit,
            ANALYSIS_TYPE.P_FCF_COMP_ANALYSIS_TYPE: self._calculate_via_comparative_p_fcf_analysis,
            ANALYSIS_TYPE.P_S_COM_ANALYSIS_TYPE: self._calculate_via_comparative_ps_analysis,
            ANALYSIS_TYPE.NET_ASSET_ANALYSIS_TYPE: self._calculate_via_net_assets,
            ANALYSIS_TYPE.DIV_ANALYSIS_TYPE: self._calculate_via_div_yeild
        }
        for (anal_mask, anal_function) in analysis_function_mapper.items():
            if self._analysis_type & anal_mask:
                anal_function()
        self._calcultate_result_price()
        
        '''
        self._calculate_via_comparative_pe_analysis()
        self._calculate_via_comparative_ev_ebit()
        self._calculate_via_comparative_p_fcf_analysis()
        self._calculate_via_comparative_ps_analysis()
        self._calculate_via_net_assets()
        '''
        
        
if __name__=="__main__":
    
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("-csv_db", dest="csv_db", type=str)
    arg_parser.add_argument("-out_file", dest="out_file", type=str)
    arg_parser.add_argument("-analysis_type", dest="analysis_type", type=str)
    args = arg_parser.parse_args()
    
    analysis_type_list = list(args.analysis_type.split(sep='|'))
    analysis_type = ANALYSIS_TYPE.NONE_ANALYSYS_TYPE
    for an_type in analysis_type_list:
        analysis_type |= ANALYSIS_TYPE[an_type]
    
    
    gen_analysis = GeneralFundamentalAnalysis(args.csv_db, analysis_type)
    gen_analysis.process()
    gen_analysis.dump_to_excel(args.out_file)
    pass
    